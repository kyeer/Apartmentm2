
import bs4
import requests
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import selenium
import time
import pandas as pd
import datetime
import smtplib
import premailer
import numpy as np
import googlemaps

def get_location1 ():
    global a
    global b
    global i
    if 'walk' in df.loc[b]['TTC or Car'] or 'ttc' in df.loc[b]['TTC or Car']:
    #if 'ttcs broken' in df.loc[b]['TTC or Car']:
        try:


            client = googlemaps.Client(key='AIzaSyCouL7oS-mK8sfx_TDdQQ8GfoU7WDeNRuQ')


            r=client.distance_matrix(a.strip(), str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
            if 'NOT_FOUND' in r['rows'][0]['elements'][0]['status']:
                if website=='craigslist':
                    r=client.distance_matrix(result2[0]['address_components'][-1]['long_name'].replace(' ',''), str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
                elif website=='kijiji':
                    r=client.distance_matrix(a.strip()[-6:], str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
                    
                    
                    
            df.loc[b]['Distance To Chosen By TTC']=r['rows'][0]['elements'][0]['duration']['value']/60*1.1
            print r['rows'][0]['elements'][0]['duration']['value']/60*1.1


            if df.loc[b]['ChosenAddress']=='nan, Ontario':
                pass
            elif int(df.loc[b]['Distance To Chosen By TTC'])<=20:
                pass
            else:
                df.loc[b]['WentWrong']= 'too far from chosen(TTC)'
                b=b+1
                i=i+1
                return


        except:
            pass



    else:
        try:
            try:
                alert = driver4.switch_to_alert()
                alert.accept()
            except:
                pass
            inputElement = driver4.find_element_by_id("Source")
            inputElement.clear()
            inputElement.send_keys(a.strip())
            inputElement = driver4.find_element_by_id("Destination")
            inputElement.clear()
            inputElement.send_keys(str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario')
            inputElement.send_keys(Keys.ENTER)
            time.sleep(2)
            html7 = driver4.page_source
            soup=bs4.BeautifulSoup(html7,'lxml')
            find11=soup.find_all('title')
            for x in find11:
                if 'hr' not in x.text[:10]:

                    if 'car' in df.loc[b]['TTC or Car']:
                        df.loc[b]['Distance To Chosen By Car']=int(x.text.strip()[:2])*1.3
                        print 'chosen' +str(int(x.text.strip()[:2])*1.3)
                        break
                    elif 'bike' in df.loc[b]['TTC or Car'] or 'ttc' in df.loc[b]['TTC or Car'] or 'walk' in df.loc[b]['TTC or Car']:
                        df.loc[b]['Distance To Chosen By Car']=int(x.text.strip()[:2])*2
                        print 'chosen' +str(int(x.text.strip()[:2])*2)
                        break
                    else:
                        df.loc[b]['Distance To Chosen By Car']=int(x.text.strip()[:2])*1.3
                        break
            if df.loc[b]['ChosenAddress']=='nan, Ontario':
                pass
            elif int(df.loc[b]['Distance To Chosen By Car'])<25:
                pass
            else:
                df.loc[b]['WentWrong']= 'too far from chosen (car)'
                b=b+1
                i=i+1
                return
        except:
            pass





    if 'walk' in df.loc[b]['TTC or Car'] or 'ttc' in df.loc[b]['TTC or Car']:
    #if 'ttcbusted' in df.loc[b]['TTC or Car']:
        try:    

            client = googlemaps.Client(key='AIzaSyCouL7oS-mK8sfx_TDdQQ8GfoU7WDeNRuQ')


            r=client.distance_matrix(a.strip(), str(results2['What is the closest intersection to work?'].iloc[number])+', Ontario',mode= 'transit')
            if 'NOT_FOUND' in r['rows'][0]['elements'][0]['status']:
                if website=='craigslist':
                    r=client.distance_matrix(result2[0]['address_components'][-1]['long_name'].replace(' ',''), str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
                elif website=='kijiji':
                    r=client.distance_matrix(a.strip()[-6:], str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
                    
            df.loc[b]['Distance To Work By TTC']=r['rows'][0]['elements'][0]['duration']['value']/60*1.1
            print r['rows'][0]['elements'][0]['duration']['value']/60*1.1

        except:
            pass

    else:
        try:
            if i>0:
                try:
                    alert = driver4.switch_to_alert()
                    alert.accept()
                except:
                    pass
            inputElement = driver4.find_element_by_id("Source")
            inputElement.clear()
            inputElement.send_keys(a.strip())
            inputElement = driver4.find_element_by_id("Destination")
            inputElement.clear()
            inputElement.send_keys(str(results2['What is the closest intersection to work?'].iloc[number])+', Toronto, Ontario')
            inputElement.send_keys(Keys.ENTER)
            time.sleep(2)
            html7 = driver4.page_source
            soup=bs4.BeautifulSoup(html7,'lxml')
            find11=soup.find_all('title')
            for x in find11:
                if 'hr' not in x.text[:10]:
                    if 'car' in df.loc[b]['TTC or Car']:
                        df.loc[b]['Distance To Work By Car']=int(x.text.strip()[:2])*1.3
                        print 'work' +str(int(x.text.strip()[:2])*1.3)
                        break
                    elif 'bike' in df.loc[b]['TTC or Car'] or 'ttc' in df.loc[b]['TTC or Car'] or 'walk' in df.loc[b]['TTC or Car']:
                        df.loc[b]['Distance To Work By Car']=int(x.text.strip()[:2])*2
                        print 'work' +str(int(x.text.strip()[:2])*2)
                        break
                    else:
                        df.loc[b]['Distance To Work By Car']=int(x.text.strip()[:2])*1.3
                        break
            else:    
                df.loc[b]['WentWrong']= 'too far from work (car)'
                b=b+1
                i=i+1
                return


        except:
            pass

    try:
        if df.loc[b]['Closer to Home or Work']=='work':
            time1=int(df.loc[b]['Current Commute Time'])
        else:
            time1=45

        df['Work Time'] = df['Distance To Work By Car'].combine_first(df['Distance To Work By TTC'])
        if df.loc[b]['WorkAddress']=='nan, Toronto, Ontario':
            pass  
        elif int(df.loc[b]['Work Time'])>time1:
            df.loc[b]['WentWrong']= 'too far from work'
            b=b+1
            i=i+1
            return
        else:
            pass
    except:
        pass




    if 'walk' in df.loc[b]['TTC or Car'] or 'ttc' in df.loc[b]['TTC or Car']:
    #if 'ttcbusted' in df.loc[b]['TTC or Car']:
        try:

            client = googlemaps.Client(key='AIzaSyCouL7oS-mK8sfx_TDdQQ8GfoU7WDeNRuQ')


            r=client.distance_matrix(a.strip(), str(results2['Outside of work, what part of the city are you usually in? '].iloc[number])+', Ontario',mode= 'transit')
            if 'NOT_FOUND' in r['rows'][0]['elements'][0]['status']:
                if website=='craigslist':
                    r=client.distance_matrix(result2[0]['address_components'][-1]['long_name'].replace(' ',''), str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
                elif website=='kijiji':
                    r=client.distance_matrix(a.strip()[-6:], str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
                    
            
            df.loc[b]['Distance To Hangout By TTC']=r['rows'][0]['elements'][0]['duration']['value']/60*1.1
            print r['rows'][0]['elements'][0]['duration']['value']/60*1.1
        except:
            pass

    else:
        try:
            try:
                alert = driver4.switch_to_alert()
                alert.accept()
            except:
                pass
            inputElement = driver4.find_element_by_id("Source")
            inputElement.clear()
            inputElement.send_keys(a.strip())
            inputElement = driver4.find_element_by_id("Destination")
            inputElement.clear()
            inputElement.send_keys(str(results2['Outside of work, what part of the city are you usually in? '].iloc[number])+', Ontario')
            inputElement.send_keys(Keys.ENTER)
            time.sleep(2)
            html7 = driver4.page_source
            soup=bs4.BeautifulSoup(html7,'lxml')
            find11=soup.find_all('title')
            for x in find11:
                if 'hr' not in x.text[:10]:
                    if 'car' in df.loc[b]['TTC or Car']:
                        df.loc[b]['Distance To Hangout By Car']=int(x.text.strip()[:2])*1.3
                        print 'hangout' +str(int(x.text.strip()[:2])*1.3)
                        break
                    elif 'bike' in df.loc[b]['TTC or Car'] or 'ttc' in df.loc[b]['TTC or Car'] or 'walk' in df.loc[b]['TTC or Car']:
                        df.loc[b]['Distance To Hangout By Car']=int(x.text.strip()[:2])*2
                        print 'hangout' +str(int(x.text.strip()[:2])*2)
                        break

                    else:
                        df.loc[b]['Distance To Hangout By Car']=int(x.text.strip()[:2])*1.3
                        break

        except:
            pass


    df['Hangout Time'] = df['Distance To Hangout By Car'].combine_first(df['Distance To Hangout By TTC'])

    try:
        if df.loc[b]['Closer to Home or Work']=='outside-of-work':
            time2=int(df.loc[b]['Current Commute Time'])
        else:
            time2=45
        if df.loc[b]['HangoutAddress']=='nan, Toronto':
            pass  
        elif int(df.loc[b]['Hangout Time'])>time2:
            df.loc[b]['WentWrong']= 'too far from hangout'
            b=b+1
            i=i+1
            return
        else:
            pass
    except:
        pass


    if 'walk' in df.loc[b]['TTC or Car'] or 'ttc' in df.loc[b]['TTC or Car']:
    #if 'ttcs busted' in df.loc[b]['TTC or Car']:
        try:

            client = googlemaps.Client(key='AIzaSyCouL7oS-mK8sfx_TDdQQ8GfoU7WDeNRuQ')


            r=client.distance_matrix(a.strip(), 'Yonge and Dundas, Toronto',mode= 'transit')
            if 'NOT_FOUND' in r['rows'][0]['elements'][0]['status']:
                if website=='craigslist':
                    r=client.distance_matrix(result2[0]['address_components'][-1]['long_name'].replace(' ',''), str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
                elif website=='kijiji':
                    r=client.distance_matrix(a.strip()[-6:], str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario',mode= 'transit')
                    
            df.loc[b]['Distance To Downtown By TTC']=r['rows'][0]['elements'][0]['duration']['value']/60*1.3*1.1
            print 'downtown'

        except:
            pass



    else:
        try:
            try:
                alert = driver4.switch_to_alert()
                alert.accept()
            except:
                pass
            inputElement = driver4.find_element_by_id("Source")
            inputElement.clear()
            inputElement.send_keys(a.strip())
            inputElement = driver4.find_element_by_id("Destination")
            inputElement.clear()
            inputElement.send_keys('Yonge and Dundas'+', Toronto')
            inputElement.send_keys(Keys.ENTER)
            time.sleep(2)
            html7 = driver4.page_source
            soup=bs4.BeautifulSoup(html7,'lxml')
            find11=soup.find_all('title')
            for x in find11:
                if 'hr' not in x.text[:10]:
                    if 'car' in df.loc[b]['TTC or Car']:
                        df.loc[b]['Distance To Downtown By Car']=int(x.text.strip()[:2])*1.3
                        print 'chosen' +str(int(x.text.strip()[:2])*1.3)
                        break
                    elif 'bike' in df.loc[b]['TTC or Car'] or 'ttc' in df.loc[b]['TTC or Car'] or 'walk' in df.loc[b]['TTC or Car']:
                        df.loc[b]['Distance To Downtown By Car']=int(x.text.strip()[:2])*2
                        print 'chosen' +str(int(x.text.strip()[:2])*2)
                        break

                    else:
                        df.loc[b]['Distance To Downtown By Car']=int(x.text.strip()[:2])*1.3
                        break

        except:
            pass

    try:
        df['Downtown Time'] = df['Distance To Downtown By Car'].combine_first(df['Distance To Downtown By TTC'])

    except:
        pass

    try:
        try:
            df['Test Time'] = df['Hangout Time'].combine_first(df['Work Time'])

        except:
            df['Chosen Time'] = df['Distance To Chosen By Car'].combine_first(df['Distance To Chosen By TTC'])

    except:     
        df.loc[b]['WentWrong']= 'could not find hangout'
        b=b+1
        i=i+1
        return


    print 'successfully completed'
    b=b+1
    i=i+1

PROXY = "fr.proxymesh.com:31280" # IP:PORT or HOST:PORT
chrome_options = Options()

chrome_options.add_experimental_option( "prefs",{'profile.managed_default_content_settings.javascript': 2,"profile.managed_default_content_settings.images":2, "profile.default_content_settings.cookies": 2})
chrome_options.add_argument('--proxy-server=%s' % PROXY)

driver6 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe")
driver6.get('http://apartmentrank.com/wp-admin')
inputElement = driver6.find_element_by_id("user_login")
inputElement.send_keys('kyeer12')
inputElement = driver6.find_element_by_id("user_pass")
inputElement.send_keys('y!5TKnJaKiJNRaL1RnN4O)or')
inputElement.send_keys(Keys.ENTER)
time.sleep(5)
driver6.get('http://apartmentrank.com/wp-admin/edit.php?post_status=all&post_type=nf_sub&form_id=3&download_all=all-subs')
link = driver6.find_element_by_link_text('Download All Submissions')
link.click()
time.sleep(10)
today=time.strftime("%Y-%m-%d")
results2=pd.read_csv('C:\Users\Kye Andreopoulus\Downloads'+r'\answer-the-questions-below-and-get-matched-to-your-ideal-apartment-or-condo-all-subs-'+today+'.csv', error_bad_lines=False)
#results2
results1=results2[(results2['When were you looking to move in?']=='Invalid date')]
results2=results2[(results2['When were you looking to move in?']!='Invalid date')&(results2['When were you looking to move in?'].str.len()==10)&(results2['When were you looking to move in?'] !='short-term')]
results2['Date Submitted']=pd.to_datetime(results2['Date Submitted'],format="%d/%m/%Y")
results1['Date Submitted']=pd.to_datetime(results1['Date Submitted'],format="%d/%m/%Y")
results2['When were you looking to move in?']=pd.to_datetime(results2['When were you looking to move in?'],format="%d/%m/%Y")
today= datetime.date(*map(int, today.split('-')))
results2=results2[results2['Date Submitted']<=today]
results2['Datediff']=today-results2['Date Submitted']
results1['Datediff']=today-results1['Date Submitted']
results1['Datediff']=results1['Datediff']/ np.timedelta64(1, 'D')
results1=results1[(results1['Datediff']<10)&(results1['Datediff']>=0)]



results2['Datediff']=results2['Datediff']/ np.timedelta64(1, 'D')
results2['Targeted']=0
results2['Targeted'][(results2['Datediff']<10)&(results2['Datediff']>=0)&(((results2['When were you looking to move in?']-results2['Date Submitted'])/ np.timedelta64(1, 'D'))<=60)] =1
results2['Targeted'][(results2['Datediff']%7==0)&(results2['Datediff']>=0)&(((results2['When were you looking to move in?']-results2['Date Submitted'])/ np.timedelta64(1, 'D'))>60)] =1
results2['Targeted'][(results2['Datediff']==1)&(((results2['When were you looking to move in?']-results2['Date Submitted'])/ np.timedelta64(1, 'D'))>60)] =1
results2['Targeted'][(results2['Datediff']==2)&(((results2['When were you looking to move in?']-results2['Date Submitted'])/ np.timedelta64(1, 'D'))>60)] =1
results2=results2[results2['Targeted']==1]
results2=results2[results2['How Many Bedrooms Were You Looking For?']!='a:0:{}']

results2=results2.append(results1)
results2=results2.reset_index()

for number in range(0,3):
    h=[]
    index =range(0, 10000)
    df=pd.DataFrame(index=index, columns=('Website','ApartmentType','WorkAddress', 'HangoutAddress', 'ChosenAddress',  
                                      'ApartmentAddress', 'Distance To Work By Car', 'Distance To Work By TTC', 
                                      'Distance To Hangout By Car', 'Distance To Hangout By TTC', 'Distance To Chosen By Car',
                                      'Distance To Chosen By TTC', 'Current Commute Time', 'Closer to Home or Work',
                                      'TTC or Car', 'Distance To Downtown By Car', 'Distance To Downtown By TTC',
                                      'Min Price', 'Max Price', 'Apartment Price', 'Basement or Above Ground', 'Pet Friendly', 'PostalCode', 'WentWrong'))


    pd106=pd.DataFrame( columns=('Apartment Location', 'Pharmacy', 'Restaurant', 'Liquor Store', 'Grocery', 'Park', 'Gym', 'CoffeeShop'))
    

    
    print 'PERSON '+str(number)
    apartmenttype=[]
    b=0
    i=0
    websitelist=['craigslist', 'kijiji' ]
    for  website in websitelist:
        print website
        if website=='kijiji':
            driver = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options)
            if 'bachelor' in results2['How Many Bedrooms Were You Looking For?'].iloc[number]:
                apartmenttype.append('http://www.kijiji.ca/b-bachelor-studio-apartments-condos/city-of-toronto/c211l1700273')
            if 'one' in results2['How Many Bedrooms Were You Looking For?'].iloc[number]:
                apartmenttype.append('http://www.kijiji.ca/b-1-bedroom-den-apartments-condos/city-of-toronto/c213l1700273')
                apartmenttype.append('http://www.kijiji.ca/b-1-bedroom-apartments-condos/city-of-toronto/c212l1700273')
            if 'two' in results2['How Many Bedrooms Were You Looking For?'].iloc[number]:
                apartmenttype.append('http://www.kijiji.ca/b-2-bedroom-apartments-condos/city-of-toronto/c214l1700273')
            if 'three' in results2['How Many Bedrooms Were You Looking For?'].iloc[number]:
                apartmenttype.append('http://www.kijiji.ca/b-3-bedroom-apartments-condos/city-of-toronto/c215l1700273')
            
            if apartmenttype==[]:
                break

            for apartmenttype2 in apartmenttype:
                url10=apartmenttype2  

                end=results2['Apartment Price Range (including utilities, parking)'].iloc[number].find('-')
                if results2['Apartment Price Range (including utilities, parking)'].iloc[number][:end]!='n':
                    minprice=results2['Apartment Price Range (including utilities, parking)'].iloc[number][:end]
                else:
                    minprice=1000
                if results2['Apartment Price Range (including utilities, parking)'].iloc[number][end+1:]!='a':
                    maxprice=results2['Apartment Price Range (including utilities, parking)'].iloc[number][end+1:]
                else:
                    maxprice=1200
                url10=url10+'?price='+str(minprice)+'__'+str(maxprice)

                url10=url10+'&ad=offering'

                #GET PETS  
                if 'no' in results2['Do you have any pets?'].iloc[number]:
                    u=0
                else:
                    url10=url10+'&pet-friendly=1'
                    u=1






                g=0


                driver2 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options)
                driver2.get('https://myttc.ca/')
                driver4 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe")
                driver4.get('http://www.distancesfrom.com/Travel-Time.aspx')




                pagenumber=0
                kijijilist=['','page-2/','page-3/','page-4/','page-5/']

                for z in kijijilist:
                    pagenumber=pagenumber+1
                    if pagenumber ==5:
                        driver.quit()
                        driver = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options)
                    brk=url10.find('c2')
                    url10=url10[0:brk]+z+url10[brk:]
                    driver.get(url10)
                    #time.sleep(5)
                    print 'PAGE ' +str(pagenumber)
                    brk2=str(driver.current_url).find('c2')
                    brk3=url10.find('c2')
                    if str(driver.current_url)[brk2-2] !=url10[brk3-2] and pagenumber>1:
                        break
                    html = driver.page_source
                    soup=bs4.BeautifulSoup(html,'lxml')
                    find=soup.find_all('div', {"class":"title"})


                    for x in find:
                        print i

                        if 'bachelor-studio' in url10:
                            df.loc[b]['ApartmentType']='Bachelor'
                        elif '1-bedroom' in url10:
                            df.loc[b]['ApartmentType']='One Bedroom'
                        elif '2-bedroom' in url10:
                            df.loc[b]['ApartmentType']='Two Bedroom'
                        elif '3-bedroom' in url10:
                            df.loc[b]['ApartmentType']='Three Bedroom'


                        #ADD PRICE
                        df.loc[b]['Min Price']=minprice
                        df.loc[b]['Max Price']=maxprice


                        #ADD PET CONFIRMATION
                        if u>0:
                            df.loc[b]['Pet Friendly']='Yes'
                        else:
                            df.loc[b]['Pet Friendly']='No'

                        df.loc[b]['Closer to Home or Work']=results2['Would you prefer to live closer to work or where you spend your free time?'].iloc[number]
                        df.loc[b]['TTC or Car']=results2['How do you usually get to work?'].iloc[number]









                        #GET MAX DISTANCE
                        for s in results2['How long is your commute?'].iloc[number].split('-'):
                            if s.isdigit():
                                if int(s.strip())==1 or int(s.strip())==45:
                                    s=45
                                else:
                                    s=int(s.strip())+15
                                df.loc[b]['Current Commute Time']= s
                                break





                        try:
                            title=x.text.strip() 
                            print title
                            link = driver.find_element_by_link_text(title)
                            link.click()
                            time.sleep(2)
                            df.loc[b]['Website']=driver.current_url
                            print driver.current_url
                            html2 = driver.page_source
                            soup2=bs4.BeautifulSoup(html2,'lxml')
                            basementcheck=soup2.find_all('meta')
                            if 'basement' in title or 'Basement' in title:
                                df.loc[b]['Basement or Above Ground']='Basement'
                                
                            else:
                                for x in basementcheck:
                                    if 'basement' in x['content'] or 'Basement' in x['content']:
                                        df.loc[b]['Basement or Above Ground']='Basement'
                                        break
                                else:
                                    df.loc[b]['Basement or Above Ground']='Above Ground'
                                    
                            if 'light' in results2['Do you prefer space or sunlight?'].iloc[number]:
                                if df.loc[b]['Basement or Above Ground']=='Basement':
                                    df.loc[b]['WentWrong']= 'Basement, No Good'
                                    driver.execute_script("window.history.go(-1)")
                                    b=b+1
                                    i=i+1
                                    continue
                                    
                                
                            
                            find2=soup2.find_all('tr')
                            j=0
                            l=0
                            for x in find2:
                                if j==1:
                                        h.append(x.text)
                                        break
                                if 'Address' in x.text.strip():
                                    j=j+1
                                    l=l+1
                                    z= x.text.strip().replace('Address','')
                                    z=str(z)
                                    a=z.strip().replace('View map','')
                                    print a.strip()
                                    if df['ApartmentAddress'].str.contains(str(a.strip())).any():
                                        df.loc[b]['WentWrong']= 'Duplicate'    
                                        break
                                    else:
                                        pass
                                    
                                    df.loc[b]['ApartmentAddress']=a.strip()
                                    
                                    print 'check1'

                                    df.loc[b]['WorkAddress']=str(results2['What is the closest intersection to work?'].iloc[number])+', Toronto, Ontario, Canada'
                                    df.loc[b]['HangoutAddress']=str(results2['Outside of work, what part of the city are you usually in? '].iloc[number])+', Ontario, Canada'
                                    try:
                                        df.loc[b]['ChosenAddress']=str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario'
                                    except:
                                        df.loc[b]['ChosenAddress']='nan, Ontario'
                                if 'Price' and '$' in x.text.strip():
                                    y=x.text.strip()
                                    yz=y.strip().replace('Price','')
                                    df.loc[b]['Apartment Price']= yz.strip()
                                    driver.execute_script("window.history.go(-1)")
                            if df.loc[b]['ApartmentAddress']!=a.strip():
                  
                                df.loc[b]['WentWrong']= 'failure to get address'
                                b=b+1
                                i=i+1
                                continue
                            else:
                                pass
                            print 'check2'
                        except:
                            #time.sleep(5)
                            
                            df.loc[b]['WentWrong']= 'failure to get address2'
                            b=b+1
                            i=i+1
                            if len(driver.current_url)>130:
                                driver.execute_script("window.history.go(-1)")
                            continue


                        print 'price successful'



                        try:
                            apartmentprice=float(df.loc[b]['Apartment Price'][1:].replace(',',''))
                        except:
                            b=b+1
                            i=i+1  
                            continue

                        print 'price transform successful'

                        if apartmentprice<=int(df.loc[b]['Max Price']) and apartmentprice>=int(df.loc[b]['Min Price']):
                            pass
                        else:
                            b=b+1
                            i=i+1
                            df.loc[b]['WentWrong']= 'price not matched'
                            continue 



                        get_location1 ()
                    
        
        elif website=='craigslist':
            driver = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options)
            if 'bachelor' in results2['How Many Bedrooms Were You Looking For?'].iloc[number]:
                apartmenttype.append('https://toronto.craigslist.ca/search/tor/apa?query=bachelor')
            if 'one' in results2['How Many Bedrooms Were You Looking For?'].iloc[number]:
                apartmenttype.append('https://toronto.craigslist.ca/search/tor/apa?query=1+bedroom')
            if 'two' in results2['How Many Bedrooms Were You Looking For?'].iloc[number]:
                apartmenttype.append('https://toronto.craigslist.ca/search/tor/apa?query=2+bedroom&bedrooms=2')
            if 'three' in results2['How Many Bedrooms Were You Looking For?'].iloc[number]:
                apartmenttype.append('https://toronto.craigslist.ca/search/tor/apa?query=3+bedroom&bedrooms=3')


            for apartmenttype2 in apartmenttype:
                url10=apartmenttype2  

                end=results2['Apartment Price Range (including utilities, parking)'].iloc[number].find('-')
                if results2['Apartment Price Range (including utilities, parking)'].iloc[number][:end]!='n':
                    minprice=results2['Apartment Price Range (including utilities, parking)'].iloc[number][:end]
                else:
                    minprice=1000
                if results2['Apartment Price Range (including utilities, parking)'].iloc[number][end+1:]!='a':
                    maxprice=results2['Apartment Price Range (including utilities, parking)'].iloc[number][end+1:]
                else:
                    maxprice=1200
                url10=url10+'&min_price='+str(minprice)
                url10=url10+'&max_price='+str(maxprice)



                #GET PETS  
                if 'no' in results2['Do you have any pets?'].iloc[number]:
                    u=0
                else:
                    url10=url10+'&pets_dog=1'
                    u=1






                url11=url10
                g=0


                driver4 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe")
                driver4.get('http://www.distancesfrom.com/Travel-Time.aspx')


                pagenumber=0
                craigslist=['','100','200','300','400']
                for z in craigslist:
                    
                    
                    pagenumber=pagenumber+1
                    #if pagenumber ==5:
                        #driver.quit()
                        #driver = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options)
                    if pagenumber==2:
                        url10=url10+'&s='+str(z)
                    elif pagenumber>2:
                        end=url10.find('&s=')
                        url10=url10[:end]+'&s='+str(z)
                    driver.get(url10)
                    #time.sleep(5)
                    print 'PAGE ' +str(pagenumber)
                    print url10

                    html = driver.page_source
                    soup=bs4.BeautifulSoup(html,'lxml')
                    
                    find=soup.find_all('a',  {"class":"result-title hdrlnk"})


                    for x in find:
                        print i

                        if 'bachelor' in url10:
                            df.loc[b]['ApartmentType']='Bachelor'
                        elif '1+bedroom' in url10:
                            df.loc[b]['ApartmentType']='One Bedroom'
                        elif '2+bedroom' in url10:
                            df.loc[b]['ApartmentType']='Two Bedroom'
                        elif '3+bedroom' in url10:
                            df.loc[b]['ApartmentType']='Three Bedroom'


                        #ADD PRICE
                        df.loc[b]['Min Price']=minprice
                        df.loc[b]['Max Price']=maxprice


                        #ADD PET CONFIRMATION
                        if u>0:
                            df.loc[b]['Pet Friendly']='Yes'
                        else:
                            df.loc[b]['Pet Friendly']='No'

                        df.loc[b]['Closer to Home or Work']=results2['Would you prefer to live closer to work or where you spend your free time?'].iloc[number]
                        df.loc[b]['TTC or Car']=results2['How do you usually get to work?'].iloc[number]





                        #GET MAX DISTANCE
                        for s in results2['How long is your commute?'].iloc[number].split('-'):
                            if s.isdigit():
                                if int(s.strip())==1 or int(s.strip())==45:
                                    s=45
                                else:
                                    s=int(s.strip())+15
                                df.loc[b]['Current Commute Time']= s
                                break




                        print url10
                        print driver.current_url

                        try:
                            a=''
                            title=x.text.strip().replace('   ',' ').replace('  ',' ')
                            print title
                            try:
                                link = driver.find_element_by_link_text(title)
                            except:
                                df.loc[b]['WentWrong']= 'No Link'
                                b=b+1
                                i=i+1
                                print 'no link'
                                continue
                            link.click()
                            time.sleep(2)
                            df.loc[b]['Website']=driver.current_url
                            print driver.current_url
                            html2 = driver.page_source
                            soup2=bs4.BeautifulSoup(html2,'lxml') 
                            basementcheck=soup2.find_all('meta')
                            if 'basement' in title or 'Basement' in title:
                                df.loc[b]['Basement or Above Ground']='Basement'
                                
                            else:
                                for x in basementcheck:
                                    if 'basement' in x['content'] or 'Basement' in x['content']:
                                        df.loc[b]['Basement or Above Ground']='Basement'
                                        break
                                else:
                                    df.loc[b]['Basement or Above Ground']='Above Ground'
                                    
                            if 'light' in results2['Do you prefer space or sunlight?'].iloc[number]:
                                if df.loc[b]['Basement or Above Ground']=='Basement':
                                    df.loc[b]['WentWrong']= 'Basement No Good'
                                    driver.execute_script("window.history.go(-1)")
                                    b=b+1
                                    i=i+1
                                    continue

                            findit=soup2.find_all('div', {'class':'mapaddress'})
                            for x in findit:
                                a= x.text
                            if a==''or 'google' in a:
                                findthis=soup2.find_all('small')
                                for address5 in findthis:
                                    a= address5.text.strip()[1:-1]
                                    break 
                            print title
                            print a.strip()
                            if df['ApartmentAddress'].str.contains(str(a.strip())).any():
                                df.loc[b]['WentWrong']= 'Duplicate'
                                b=b+1
                                i=i+1
                                print driver.current_url
                                if len(driver.current_url)<80:    
                                    driver.execute_script("window.history.go(-1)")
                                continue

                            else:
                                pass




                            

                            gmaps = googlemaps.Client(key='AIzaSyCouL7oS-mK8sfx_TDdQQ8GfoU7WDeNRuQ')
                            result = gmaps.geocode(a.strip()+'Ontario, Canada')

                            result2 = gmaps.reverse_geocode( (result[0]['geometry']['location']['lat'],result[0]['geometry']['location']['lng']))
                            a=a.strip()+', '+ result2[0]['address_components'][-1]['long_name'].replace(' ','')
                            print 'check1'

                            df.loc[b]['ApartmentAddress']=a.strip()
                            df.loc[b]['WorkAddress']=str(results2['What is the closest intersection to work?'].iloc[number])+', Toronto, Ontario, Canada'
                            df.loc[b]['HangoutAddress']=str(results2['Outside of work, what part of the city are you usually in? '].iloc[number])+', Ontario, Canada'
                            try:
                                df.loc[b]['ChosenAddress']=str(results2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[number])+', Ontario'
                            except:
                                df.loc[b]['ChosenAddress']='nan, Ontario' 


                            findit=soup2.find_all('span', {'class':'price'})
                            for x in findit:
                                yz= x.text

                            df.loc[b]['Apartment Price']= yz.strip() 
                            driver.execute_script("window.history.go(-1)")


                            if df.loc[b]['ApartmentAddress']!=a.strip():
                                print df.loc[b]['ApartmentAddress']
                                print a.strip()

                                df.loc[b]['WentWrong']= 'failure to get address'
                                b=b+1
                                i=i+1
                                continue
                            else:
                                pass
                            print 'check2'
                        except:
                        
                            #time.sleep(5
                            print 'failed'
                            print len(driver.current_url)
                            print driver.current_url
                            df.loc[b]['WentWrong']= 'failure to get address2'
                            if len(driver.current_url)<80:
                                driver.execute_script("window.history.go(-1)")
                            b=b+1
                            i=i+1
                            continue


                        print 'price successful'
                        try:
                            apartmentprice=float(df.loc[b]['Apartment Price'][1:].replace(',',''))
                        except:
                            b=b+1
                            i=i+1  
                            continue


                        print 'price transform successful'

                        if apartmentprice<=int(df.loc[b]['Max Price']) and apartmentprice>=int(df.loc[b]['Min Price']):
                            pass
                        else:
                            b=b+1
                            i=i+1
                            df.loc[b]['WentWrong']= 'price not matched'
                            continue 


                        get_location1()

    dffake=df  
    df=df[df['ApartmentAddress'].notnull()]  
    df=df[df['Website'].notnull()] 
    df=df[df['WentWrong'].isnull()]
    df=df[df['Work Time'].notnull()]
    df=df[df['Hangout Time'].notnull()]
    
    if len(df)==0:
        HEADER = '''
        <html>
            <head>

            </head>
            <body>
        '''
        FOOTER = '''
            </body>
        </html>
        '''

        data=HEADER + dffake.to_html(classes='df').encode('utf8') + FOOTER

        import smtplib

        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        # me == my email address
        # you == recipient's email address
        me = "torontoapartmentrank@gmail.com"
        you = "torontoapartmentrank@gmail.com "

        # Create message container - the correct MIME type is multipart/alternative.
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'No Results for ' +results2['First Name'].iloc[number]
        msg['From'] = me
        msg['To'] = you


        # Create the body of the message (a plain-text and an HTML version).
        text = "Hi!\nHow are you?\nHere is the link you wanted:\nhttp://www.python.org"
        html = data

        # Record the MIME types of both parts - text/plain and text/html.
        part1 = MIMEText(text, 'plain')
        part2 = MIMEText(html, 'html')

        # Attach parts into message container.
        # According to RFC 2046, the last part of a multipart message, in this case
        # the HTML message, is best and preferred.
        msg.attach(part1)
        msg.attach(part2)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login("torontoapartmentrank@gmail.com", "Milou511!")

        server.sendmail("torontoapartmentrank@gmail.com", "torontoapartmentrank@gmail.com", msg.as_string())
        # sendmail function takes 3 arguments: sender's addrwalkess, recipient's address
        # and message to send - here it is sent as one string.
        server.quit()
        continue
    
    try:
        df['Work Time']=df['Work Time'].fillna(50)
        df['Hangout Time']=df['Hangout Time'].fillna(50)
        df['Work Time']=df['Work Time'].fillna(50)
        df['Hangout Time']=df['Hangout Time'].fillna(50)

    except:
        pass

    df['Chosen Time']=df['Distance To Chosen By Car'].combine_first(df['Distance To Chosen By TTC'])
    
    
    df['Total Distance']=0
    try:
        df['Work Time']=pd.to_numeric(df['Work Time'])
    except:
        print'No Work'
        try:
            df['Hangout Time']=pd.to_numeric(df['Hangout Time'])
            df['Total Distance']=df['Hangout Time']
        except:
            print 'No Hangout'
            df['Chosen Time']=pd.to_numeric(df['Chosen Time'])
            df['Total Distance']=df['Chosen Time']
        pass
    try:
        df['Hangout Time']=pd.to_numeric(df['Hangout Time'])
    except:
        pass
    try:
        df['Total Distance'][(df['Closer to Home or Work']=='work')&(df['Hangout Time']<df['Work Time'])]= df['Work Time']+df['Work Time']+5
        df['Total Distance'][(df['Closer to Home or Work']=='outside-of-work')&(df['Hangout Time']>df['Work Time'])]= df['Hangout Time']+df['Hangout Time']+5
        df['Total Distance'][df['Total Distance']==0]=df['Work Time']+df['Hangout Time']
    except:
        pass

    df = df.sort(['Total Distance'])

    if len(df.index)>15:
        df=df.head(15)
    else:
        pass
    
    df=df[df['Website'].notnull()]        
    df=df[df['WentWrong'].isnull()]
    
    deleteafter=[]


    for x in range(0,len(df.index)):
        #Liquor Store
        
    
        numberx=x
        driver13 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options) 
        location=df['ApartmentAddress'].iloc[x].replace(" ", "+")
        location=location.replace('&', ' ')



        driver13.get('https://www.yelp.com/search?find_desc=liquor+store&find_loc='+location)
        #time.sleep(5)
        html13 = driver13.page_source
        soup13=bs4.BeautifulSoup(html13,'lxml')



        index =range(0, 10)
        dfYelp=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Liquor Store'))

        num=0


        try:
            find3=soup13.find_all('div')
            for x in find3:
                try:
                    if 'star rating' in x['title']:
                        dfYelp.loc[num]['Stars']=str(x['title'][:3])
                        dfYelp.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                except:
                    continue

            num=0
            find1=soup13.find_all('small')
            for x in find1:
                if num ==10:
                    break
                dfYelp.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0
            find2=soup13.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==10:
                    break
                dfYelp.loc[num]['Liquor Store']=x.text.strip()
                num=num+1

            dfYelp=dfYelp[dfYelp.notnull()]
            dfYelp1=dfYelp[dfYelp['Liquor Store'].str.contains('LCBO')]
            dfYelp2=dfYelp[dfYelp['Liquor Store'].str.contains('Beer Store')]
            dfYelp3=dfYelp[dfYelp['Liquor Store'].str.contains('Wine Rack')]


            frames = [dfYelp1, dfYelp2, dfYelp3]
            result = pd.concat(frames)

            result['Distance']=pd.to_numeric(result['Distance'])

            liquorresult=result[result['Distance']<=1]

            try:
                pd4=liquorresult.ix[liquorresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd5=pd4.transpose()

            except:
                pd5=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Liquor Store'))
                pass

        except:
            l=0
            link = driver13.find_elements_by_xpath('//*[@href]')
            tobreak=0
            for y in link:
                l=l+1
                if l==30:
                    try:
                        driver13.get(y.get_attribute('href'))
                    except:
                        deleteafter.append(df['ApartmentAddress'].iloc[x])
                        tobreak=1
                    break
            if tobreak==1:
                continue
            html13 = driver13.page_source
            soup13=bs4.BeautifulSoup(html13,'lxml')
            find3=soup13.find_all('div')
            for x in find3:
                try:
                    if 'star rating' in x['title']:
                        dfYelp.loc[num]['Stars']=str(x['title'][:3])
                        dfYelp.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                except:
                    continue

            num=0
            find1=soup13.find_all('small')
            for x in find1:
                if num ==3:
                    break
                dfYelp.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0

            find2=soup13.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==3:
                    break
                dfYelp.loc[num]['Liquor Store']=x.text.strip()

                num=num+1

            dfYelp=dfYelp[dfYelp['Liquor Store'].notnull()]
            dfYelp1=dfYelp[dfYelp['Liquor Store'].str.contains('LCBO')]
            dfYelp2=dfYelp[dfYelp['Liquor Store'].str.contains('Beer Store')]
            dfYelp3=dfYelp[dfYelp['Liquor Store'].str.contains('Wine Rack')]


            frames = [dfYelp1, dfYelp2, dfYelp3]
            liquorresult = pd.concat(frames)

            try:
                pd4=liquorresult.ix[liquorresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd5=pd4.transpose()

            except:
                pd5=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Liquor Store'))
                pass





        #Restaurant

        driver14 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options) 
        #location=a.strip().replace(" ", "+")


        driver14.get('https://www.yelp.com/search?find_desc=restaurant&find_loc='+location)
        #time.sleep(5)
        html14 = driver14.page_source
        soup14=bs4.BeautifulSoup(html14,'lxml')



        index =range(0, 10)
        dfRestaurant=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Restaurant'))

        num=0


        try:
            find4=soup14.find_all('div')
            for x in find4:
                try:
                    if 'star rating' in x['title']:
                        dfRestaurant.loc[num]['Stars']=str(x['title'][:3])
                        dfRestaurant.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                except:
                    continue

            num=0
            find1=soup14.find_all('small')
            for x in find1:
                if num ==10:
                    break
                dfRestaurant.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0
            find2=soup14.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==10:
                    break
                dfRestaurant.loc[num]['Restaurant']=x.text.strip()
                num=num+1

            result=dfRestaurant[dfRestaurant.notnull()]




            result['Distance']=pd.to_numeric(result['Distance'])

            restaurantresult=result[result['Distance']<=1]
            try:
                pd4=restaurantresult.ix[restaurantresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd6=pd4.transpose()
            except:
                pd6=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Restaurant'))
                pass
        except:
            l=0
            link = driver14.find_elements_by_xpath('//*[@href]')
            for x in link:
                l=l+1
                if l==30:
                    driver14.get(x.get_attribute('href'))
                    break
            html14 = driver14.page_source
            soup14=bs4.BeautifulSoup(html14,'lxml')
            find4=soup14.find_all('div')
            for x in find3:
                try:
                    if 'star rating' in x['title']:
                        dfRestaurant.loc[num]['Stars']=str(x['title'][:3])
                        dfRestaurant.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                except:
                    continue

            num=0
            find1=soup14.find_all('small')
            for x in find1:
                if num ==3:
                    break
                dfRestaurant.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0

            find2=soup14.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==3:
                    break
                dfRestaurant.loc[num]['Restaurant']=x.text.strip()

                num=num+1

            restaurantresult=dfRestaurant[dfRestaurant['Restaurant'].notnull()]



            try:
                pd4=restaurantresult.ix[restaurantresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd6=pd4.transpose()
            except:
                pd6=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Restaurant'))
                pass



        driver15 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options) 
        #location=a.strip().replace(" ", "+")


        driver15.get('https://www.yelp.com/search?find_desc=pharmacy&find_loc='+location)
        #time.sleep(5)
        html14 = driver15.page_source
        soup14=bs4.BeautifulSoup(html14,'lxml')  


        index =range(0, 10)
        dfpharmacy=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Pharmacy'))

        num=0


        try:
            find4=soup14.find_all('div')
            for x in find4:
                try:
                    if 'star rating' in x['title']:
                        dfpharmacy.loc[num]['Stars']=str(x['title'][:3])
                        dfpharmacy.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                except:
                    continue

            num=0
            find1=soup14.find_all('small')
            for x in find1:
                if num ==10:
                    break
                dfpharmacy.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0
            find2=soup14.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==10:
                    break
                dfpharmacy.loc[num]['Pharmacy']=x.text.strip()
                num=num+1

            result=dfpharmacy[dfpharmacy.notnull()]




            result['Distance']=pd.to_numeric(result['Distance'])

            pharmacyresult=result[result['Distance']<=1]

            try:

                pd4=pharmacyresult.ix[pharmacyresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd7=pd4.transpose()
            except:
                pd7=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Pharmacy'))
                pass



        except:
            l=0
            link = driver15.find_elements_by_xpath('//*[@href]')
            for x in link:
                l=l+1
                if l==30:
                    driver15.get(x.get_attribute('href'))
                    break
            html14 = driver15.page_source
            soup14=bs4.BeautifulSoup(html14,'lxml')
            find4=soup14.find_all('div')
            for x in find3:
                try:
                    if 'star rating' in x['title']:
                        dfpharmacy.loc[num]['Stars']=str(x['title'][:3])
                        dfpharmacy.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                except:
                    continue

            num=0
            find1=soup14.find_all('small')
            for x in find1:
                if num ==3:
                    break
                dfpharmacy.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0

            find2=soup14.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==3:
                    break
                dfpharmacy.loc[num]['Pharmacy']=x.text.strip()

                num=num+1

            pharmacyresult=dfpharmacy[dfpharmacy['Pharmacy'].notnull()]



            try:

                pd4=pharmacyresult.ix[pharmacyresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd7=pd4.transpose()
            except:
                pd7=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Pharmacy'))
                pass



        pd100=pd.merge(pd7,pd6, left_on='Apartment Location', right_on='Apartment Location', how='left')
        pd101=pd.merge(pd100,pd5, left_on='Apartment Location', right_on='Apartment Location', how='left')

        

        driver16 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options) 
        #location=a.strip().replace(" ", "+")


        driver16.get('https://www.yelp.com/search?find_desc=grocery&find_loc='+location)
        #time.sleep(5)
        html15 = driver16.page_source
        soup15=bs4.BeautifulSoup(html15,'lxml')  


        index =range(0, 10)
        dfGrocery=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Grocery'))

        num=0


        try:
            find4=soup15.find_all('div')
            for x in find4:
                try:
                    if 'star rating' in x['title']:
                        dfGrocery.loc[num]['Stars']=str(x['title'][:3])
                        dfGrocery.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                        if num==3:
                            break
                except:
                    continue

            num=0
            find1=soup15.find_all('small')
            for x in find1:
                if num ==5:
                    break
                dfGrocery.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0
            find2=soup15.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==5:
                    break
                dfGrocery.loc[num]['Grocery']=x.text.strip()
                num=num+1

            result=dfGrocery[dfGrocery.notnull()]




            result['Distance']=pd.to_numeric(result['Distance'])

            groceryresult=result[result['Distance']<=1]

            try:
                pd4=groceryresult.ix[groceryresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd8=pd4.transpose()

            except:
                pd8=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Grocery'))
                pass



        except:
            l=0
            link = driver16.find_elements_by_xpath('//*[@href]')
            for x in link:
                l=l+1
                if l==30:
                    driver16.get(x.get_attribute('href'))
                    break
            html15 = driver16.page_source
            soup15=bs4.BeautifulSoup(html15,'lxml')
            find5=soup14.find_all('div')
            for x in find5:
                try:
                    if 'star rating' in x['title']:
                        dfGrocery.loc[num]['Stars']=str(x['title'][:3])
                        dfGrocery.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                    if num==3:
                        break
                except:
                    continue

            num=0
            find1=soup15.find_all('small')
            for x in find1:
                if num >=3:
                    break
                dfGrocery.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0

            find2=soup15.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num >=3:
                    break
                dfGrocery.loc[num]['Grocery']=x.text.strip()
                num=num+1

            groceryresult=dfGrocery[dfGrocery['Grocery'].notnull()]



            try:
                pd4=groceryresult.ix[groceryresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd8=pd4.transpose()

            except:
                pd8=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Grocery'))
                pass

        pd102=pd.merge(pd101,pd8, left_on='Apartment Location', right_on='Apartment Location', how='left')

        

        driver17 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options) 
        #location=a.strip().replace(" ", "+")


        driver17.get('https://www.yelp.com/search?find_desc=park&find_loc='+location)
        #time.sleep(5)
        html16 = driver17.page_source
        soup16=bs4.BeautifulSoup(html16,'lxml')  


        index =range(0, 10)
        dfPark=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Park'))

        num=0


        try:
            find4=soup16.find_all('div')
            for x in find4:
                try:
                    if 'star rating' in x['title']:
                        dfPark.loc[num]['Stars']=str(x['title'][:3])
                        dfPark.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                        if num==3:
                            break
                except:
                    continue

            num=0
            find1=soup16.find_all('small')
            for x in find1:
                if num ==5:
                    break
                dfPark.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0
            find2=soup16.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==5:
                    break
                dfPark.loc[num]['Park']=x.text.strip()
                num=num+1

            result=dfPark[dfPark.notnull()]




            result['Distance']=pd.to_numeric(result['Distance'])

            parkresult=result[result['Distance']<=1]


            try:
                pd4=parkresult.ix[parkresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd9=pd4.transpose()

            except:
                pd9=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Park'))
                pass



        except:
            l=0
            link = driver17.find_elements_by_xpath('//*[@href]')
            for x in link:
                l=l+1
                if l==30:
                    driver17.get(x.get_attribute('href'))
                    break
            html16 = driver17.page_source
            soup16=bs4.BeautifulSoup(html16,'lxml')
            find5=soup14.find_all('div')
            for x in find5:
                try:
                    if 'star rating' in x['title']:
                        dfPark.loc[num]['Stars']=str(x['title'][:3])
                        dfPark.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                    if num==3:
                        break
                except:
                    continue

            num=0
            find1=soup16.find_all('small')
            for x in find1:
                if num >=3:
                    break
                dfPark.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0

            find2=soup16.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num >=3:
                    break
                dfPark.loc[num]['Park']=x.text.strip()
                num=num+1

            parkresult=dfPark[dfPark['Park'].notnull()]



            try:
                pd4=parkresult.ix[parkresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd9=pd4.transpose()

            except:
                pd9=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Park'))
                pass

        pd103=pd.merge(pd102,pd9, left_on='Apartment Location', right_on='Apartment Location', how='left')



        driver18 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options) 
        #location=a.strip().replace(" ", "+")


        driver18.get('https://www.yelp.com/search?find_desc=gym&find_loc='+location)
        #time.sleep(5)
        html16 = driver18.page_source
        soup16=bs4.BeautifulSoup(html16,'lxml')  


        index =range(0, 10)
        dfGym=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Gym'))

        num=0


        try:
            find4=soup16.find_all('div')
            for x in find4:
                try:
                    if 'star rating' in x['title']:
                        dfGym.loc[num]['Stars']=str(x['title'][:3])
                        dfGym.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                        if num==5:
                            break
                except:
                    continue

            num=0
            find1=soup16.find_all('small')
            for x in find1:
                if num ==5:
                    break
                dfGym.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0
            find2=soup16.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==5:
                    break
                dfGym.loc[num]['Gym']=x.text.strip()
                num=num+1

            result=dfGym[dfGym.notnull()]




            result['Distance']=pd.to_numeric(result['Distance'])

            gymresult=result[result['Distance']<=1]

            try:
                pd4=gymresult.ix[gymresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd10=pd4.transpose()

            except:
                pd10=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Gym'))
                pass


        except:
            l=0
            link = driver18.find_elements_by_xpath('//*[@href]')
            for x in link:
                l=l+1
                if l==30:
                    driver18.get(x.get_attribute('href'))
                    break
            html16 = driver18.page_source
            soup16=bs4.BeautifulSoup(html16,'lxml')
            find5=soup14.find_all('div')
            for x in find5:
                try:
                    if 'star rating' in x['title']:
                        dfGym.loc[num]['Stars']=str(x['title'][:3])
                        dfGym.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                    if num==3:
                        break
                except:
                    continue

            num=0
            find1=soup16.find_all('small')
            for x in find1:
                if num >=3:
                    break
                dfGym.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0

            find2=soup16.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num >=3:
                    break
                dfGym.loc[num]['Gym']=x.text.strip()
                num=num+1

            gymresult=dfGym[dfGym['Gym'].notnull()]



            try:
                pd4=gymresult.ix[gymresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd10=pd4.transpose()

            except:
                pd10=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'Gym'))
                pass


        pd104=pd.merge(pd103,pd10, left_on='Apartment Location', right_on='Apartment Location', how='left')




        driver19 = webdriver.Chrome("C:\Users\Kye Andreopoulus\Desktop\chromedriver.exe",chrome_options=chrome_options) 
        #location=a.strip().replace(" ", "+")


        driver19.get('https://www.yelp.com/search?find_desc=CoffeeShop&find_loc='+location)
        #time.sleep(5)
        html16 = driver19.page_source
        soup16=bs4.BeautifulSoup(html16,'lxml')  


        index =range(0, 10)
        dfCoffeeShop=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'CoffeeShop'))

        num=0


        try:
            find4=soup16.find_all('div')
            for x in find4:
                try:
                    if 'star rating' in x['title']:
                        dfCoffeeShop.loc[num]['Stars']=str(x['title'][:3])
                        dfCoffeeShop.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                        if num==3:
                            break
                except:
                    continue

            num=0
            find1=soup16.find_all('small')
            for x in find1:
                if num ==5:
                    break
                dfCoffeeShop.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0
            find2=soup16.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num ==5:
                    break
                dfCoffeeShop.loc[num]['CoffeeShop']=x.text.strip()
                num=num+1

            result=dfCoffeeShop[dfCoffeeShop.notnull()]




            result['Distance']=pd.to_numeric(result['Distance'])

            coffeeresult=result[result['Distance']<=1]

            try:
                pd4=coffeeresult.ix[coffeeresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd11=pd4.transpose()
            except:
                pd11=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'CoffeeShop'))
                pass


        except:
            l=0
            link = driver19.find_elements_by_xpath('//*[@href]')
            for x in link:
                l=l+1
                if l==30:
                    driver19.get(x.get_attribute('href'))
                    break
            html16 = driver19.page_source
            soup16=bs4.BeautifulSoup(html16,'lxml')
            find5=soup14.find_all('div')
            for x in find5:
                try:
                    if 'star rating' in x['title']:
                        dfCoffeeShop.loc[num]['Stars']=str(x['title'][:3])
                        dfCoffeeShop.loc[num]['Apartment Location']=location
                        location=location.replace('subway','')
                        location=location.replace('Subway','')
                        location=location.replace('SUBWAY','')
                        num=num+1
                    if num==3:
                        break
                except:
                    continue

            num=0
            find1=soup16.find_all('small')
            for x in find1:
                if num >=3:
                    break
                dfCoffeeShop.loc[num]['Distance']=str(x.text.strip()[:3])
                num=num+1

            num=0

            find2=soup16.find_all('a',{'biz-name js-analytics-click'})
            for x in find2:
                if num >=3:
                    break
                dfCoffeeShop.loc[num]['CoffeeShop']=x.text.strip()
                num=num+1

            coffeeresult=dfCoffeeShop[dfCoffeeShop['CoffeeShop'].notnull()]



            try:
                pd4=coffeeresult.ix[coffeeresult['Stars'].idxmax()]
                pd4=pd.DataFrame(pd4)

                pd11=pd4.transpose()
            except:
                pd11=pd.DataFrame(index=index, columns=('Apartment Location','Stars', 'Distance', 'CoffeeShop'))
                pass

        pd105=pd.merge(pd104,pd11, left_on='Apartment Location', right_on='Apartment Location', how='left')

        pd106=pd106.append(pd105[['Apartment Location', 'Pharmacy', 'Restaurant', 'Liquor Store', 'Grocery', 'Park', 'Gym', 'CoffeeShop']])

    
    for x in deleteafter:
        df=df[df['ApartmentAddress']!=x]
        
    
    df=df.reset_index()
    pd106=pd106.reset_index()
    df['Test']=df['ApartmentAddress'].str.replace(" ", "+")
    df['Test']=df['Test'].astype(str)
    pd106['Apartment Location']=pd106['Apartment Location'].astype(str)
    pd107=pd.merge (pd106, df,left_on='Apartment Location', right_on='Test', how='inner')








    #GET POSTALCODE
    df10=pd.read_excel('C:/Users/Kye Andreopoulus/Documents/Environics.xlsx')
    df10['average_age']='Avg. age in neighbourhood: '+ str(int(df10.median_age_of_total_population.mean()))
    df10['children']=df10.average_children_per_census_family.mean()
    df10['children_index']=(((df10.average_children_per_census_family/df10.children)-1)*100)
    df11=df10
    df11.children_index=df11.children_index.astype(int)
    pd107['PostalCode']=pd107['ApartmentAddress'].str.replace(' ','')
    pd107['PostalCode2']=pd107['PostalCode'].str[-6:]
    pd107['PostalCode2'][pd107['PostalCode2']=='Canada'] =pd107['PostalCode'].str[-13:-7]
    dffinal=pd.merge(pd107,df11, how='left', left_on='PostalCode2', right_on='postal_code')
    
    dffinal2=dffinal
    results3=results2.iloc[[number]]
    results3=results3.reset_index()
    dffinal2['children_index']=dffinal2['children_index'].fillna('0')
    dffinal2['For whom?']='Families and Singles'
    dffinal2['For whom?'][dffinal2['children_index']<=0]='Mostly Singles'
    dffinal2['For whom?'][dffinal2['children_index']>0]='Mostly Families'
    dffinal2.rename(columns = {'ApartmentType':'How Many Bedrooms Were You Looking For?'}, inplace = True)
    dffinal2.rename(columns = {'Pharmacy':'Pharmacy Within 10 Mins'}, inplace = True)
    dffinal2.rename(columns = {'Restaurant':'Restaurant Within 10 Mins'}, inplace = True)
    dffinal2.rename(columns = {'Liquor Store':'Liquor Store Within 10 Mins'}, inplace = True)
    dffinal2.rename(columns = {'Grocery':'Grocery Within 10 Mins'}, inplace = True)
    dffinal2.rename(columns = {'Park':'Park Within 10 Mins'}, inplace = True)
    dffinal2.rename(columns = {'Gym':'Gym Within 10 Mins'}, inplace = True)
    dffinal2.rename(columns = {'CoffeeShop':'Coffee Shop Within 10 Mins'}, inplace = True)
    dffinal2.rename(columns = {'median_age_of_total_population':'What age are you, in spirit?'}, inplace = True)
    dffinal2.rename(columns = {'Apartment Price':'Apartment Price Range (including utilities, parking)'}, inplace = True)
    if 'wine' in results3['At dinner you usually drink...'][0]:
        dffinal2['At dinner you usually drink...']=dffinal2['Liquor Store Within 10 Mins'] +' within 10 mins'


    dffinal2.rename(columns = {'Pet Friendly':'Do you have any pets?'}, inplace = True)

    if 'cook' in results3['Do you prefer to cook at home or eat out?'][0]:
        dffinal2['Do you prefer to cook at home or eat out?']=dffinal2['Grocery Within 10 Mins'] +' within 10 mins'
    else:
        dffinal2['Do you prefer to cook at home or eat out?']=dffinal2['Restaurant Within 10 Mins'] +' within 10 mins'

    if 'town' in  results3['After work you usually...'][0]:
        dffinal2['After work you usually...']=str(dffinal2['Downtown Time']) +' mins to Downtown'
    elif 'Park' in  results3['After work you usually...'][0]:
        dffinal2['After work you usually...']=str(dffinal2['Park Within 10 Mins']) +' within 10 mins'
    elif 'Gym' in  results3['After work you usually...'][0]:
        dffinal2['After work you usually...']=str(dffinal2['Gym Within 10 Mins']) +' within 10 mins'
    else:
        del results3['After work you usually...']
    dffinal2['What age are you, in spirit?']=dffinal2['What age are you, in spirit?'].fillna(32)
    dffinal2['What age are you, in spirit?']='  '+dffinal2['What age are you, in spirit?'].astype(int).astype(str)
    dffinal2.rename(columns = {'median_household_incomeconstant_year_2005_dollars':'Average Income of Neighbourhood'}, inplace = True)
    dffinal2['Chosen Time']=dffinal2['Chosen Time'].fillna(15)
    if 'walk' not in dffinal2['TTC or Car'][0]:
        dffinal2['What is the closest intersection to work?']= dffinal2['Work Time'].astype(int).astype(str) + ' mins by ' + dffinal2['TTC or Car']
        dffinal2['Outside of work, what part of the city are you usually in? ']= dffinal2['Hangout Time'].astype(int).astype(str) + ' mins by ' + dffinal2['TTC or Car']

        dffinal2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)']= dffinal2['Chosen Time'].astype(int).astype(str) + ' mins by ' + dffinal2['TTC or Car']
    else:

        dffinal2['What is the closest intersection to work?']= dffinal2['Work Time'].astype(int).astype(str) + ' mins by ' + 'TTC'
        dffinal2['Outside of work, what part of the city are you usually in? ']= dffinal2['Hangout Time'].astype(int).astype(str) + ' mins by ' + 'TTC'

        dffinal2['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)']= dffinal2['Chosen Time'].astype(int).astype(str) + ' mins by ' + 'TTC'
    if 'no' not in results3['Do you have any pets?'][0]:
        dffinal2['Do you have any pets?']='Pet Friendly'
        
    dffinal2['Neighbourhood']=''
    dffinal2['Bed Bug Reports in Last 2 Years']=''
    dffinal2['Safety of Neighbourhood']=''
    dffinal2['Walkscore']=''
    dffinal2['Photos']=''
    dffinal2['Are you looking for a new place to rent or own?']='Rent'
    finalresult = pd.concat([results3, dffinal2], ignore_index=True)

    if 'wine' in results3['At dinner you usually drink...'][0] and ('town' in  results2['After work you usually...'][number] or 'Park' in  results2['After work you usually...'][number] or 'Gym' in  results2['After work you usually...'][number]):
        dffinal3=finalresult[['Basement or Above Ground', 'Photos', 'ApartmentAddress', 'Apartment Price Range (including utilities, parking)','How Many Bedrooms Were You Looking For?', 'Bed Bug Reports in Last 2 Years', 'What is the closest intersection to work?','Outside of work, what part of the city are you usually in? ',  'If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)',  'Do you have any pets?', 'Do you prefer space or sunlight?',  'Are you looking for short-term or long-term?',  'Do you prefer to cook at home or eat out?', 'After work you usually...',   'At dinner you usually drink...', 'Neighbourhood', 'For whom?', 'What age are you, in spirit?', 'Safety of Neighbourhood', 'Average Income of Neighbourhood',  'Gym Within 10 Mins','Pharmacy Within 10 Mins','Restaurant Within 10 Mins', 'Grocery Within 10 Mins', 'Park Within 10 Mins','Coffee Shop Within 10 Mins','Walkscore','When were you looking to move in?','Website']]
    elif 'wine' not in results3['At dinner you usually drink...'][0] and ('town' in  results2['After work you usually...'][number] or 'Park' in  results2['After work you usually...'][number] or 'Gym' in  results2['After work you usually...'][number]):
        dffinal3=finalresult[['Basement or Above Ground','Photos', 'ApartmentAddress','Apartment Price Range (including utilities, parking)','How Many Bedrooms Were You Looking For?', 'Bed Bug Reports in Last 2 Years',  'What is the closest intersection to work?','Outside of work, what part of the city are you usually in? ',  'If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)', 'Do you have any pets?', 'Do you prefer space or sunlight?',  'Are you looking for short-term or long-term?',  'Do you prefer to cook at home or eat out?', 'After work you usually...',    'Neighbourhood', 'For whom?', 'What age are you, in spirit?', 'Safety of Neighbourhood', 'Average Income of Neighbourhood',  'Gym Within 10 Mins','Pharmacy Within 10 Mins','Restaurant Within 10 Mins', 'Grocery Within 10 Mins', 'Park Within 10 Mins','Coffee Shop Within 10 Mins', 'Walkscore','When were you looking to move in?','Website']]
    elif 'wine' in results3['At dinner you usually drink...'][0] and 'town' not in  results2['After work you usually...'][number] and 'Park' not in  results2['After work you usually...'][number] and 'Gym' not in  results2['After work you usually...'][number]:  
        dffinal3=finalresult[['Basement or Above Ground','Photos', 'ApartmentAddress','Apartment Price Range (including utilities, parking)','How Many Bedrooms Were You Looking For?', 'Bed Bug Reports in Last 2 Years', 'What is the closest intersection to work?','Outside of work, what part of the city are you usually in? ',  'If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)',  'Do you have any pets?', 'Do you prefer space or sunlight?',  'Are you looking for short-term or long-term?',  'Do you prefer to cook at home or eat out?',  'At dinner you usually drink...',   'Neighbourhood', 'For whom?', 'What age are you, in spirit?', 'Safety of Neighbourhood', 'Average Income of Neighbourhood',   'Gym Within 10 Mins','Pharmacy Within 10 Mins','Restaurant Within 10 Mins', 'Grocery Within 10 Mins', 'Park Within 10 Mins','Coffee Shop Within 10 Mins', 'Walkscore','When were you looking to move in?','Website']]
    elif 'wine' not in results3['At dinner you usually drink...'][0] and 'town' not in  results2['After work you usually...'][number] and 'Park' not in  results2['After work you usually...'][number] and 'Gym' not in  results2['After work you usually...'][number]:  
        dffinal3=finalresult[['Basement or Above Ground','Photos', 'ApartmentAddress','Apartment Price Range (including utilities, parking)','How Many Bedrooms Were You Looking For?','Bed Bug Reports in Last 2 Years', 'What is the closest intersection to work?','Outside of work, what part of the city are you usually in? ',  'If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)',  'Do you have any pets?', 'Do you prefer space or sunlight?',  'Are you looking for short-term or long-term?',  'Do you prefer to cook at home or eat out?', 'Neighbourhood', 'For whom?', 'What age are you, in spirit?', 'Safety of Neighbourhood', 'Average Income of Neighbourhood',    'Gym Within 10 Mins','Pharmacy Within 10 Mins','Restaurant Within 10 Mins', 'Grocery Within 10 Mins', 'Park Within 10 Mins','Coffee Shop Within 10 Mins','Walkscore','When were you looking to move in?','Website']]
    results3['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)']=results3['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].fillna('0')
    abba= results3['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)'].iloc[[0]]
    if abba[0]=='0':
        del dffinal3['If you know the neighbourhood you want to live in please enter it below (Note we will only search apartments in this neighbourhood)']
    dffinal3['Average Income of Neighbourhood']='$'+' '+dffinal3['Average Income of Neighbourhood'].astype(str)
    
    if 'no' in results3['Do you have any pets?'][0]:
        del dffinal3['Do you have any pets?']
        
    dffinal4=dffinal3.transpose()
    finalcolumns=['Your Answers']
    for this in range(0,len(dffinal4.columns)-1):
        add='Apartment Match '+str(this+1)
        finalcolumns.append(add)
    dffinal4.columns = finalcolumns
    dffinal4=dffinal4.fillna('')

    pd.set_option('display.max_colwidth', -1)


    HEADER = '''
    <html>
        <head>

        </head>
        <body>
    '''
    FOOTER = '''
        </body>
    </html>
    '''
    
    import requests
    import bs4
    
    secondnumber=1
    numbera=1
    locationlist=''
    laptopimage1=''
    image=''
    nextimage=''
    nextnextimage=''
    imagecount=0
    img1=''
    img2=''
    img3=''
    msgImage=''
    msgImage2=''
    msgImage3=''
    match1=''
    match2=''
    match3=''
    goodmatch=[]
    for images in range(0,len(dffinal4.columns)):
        tobreak=0
        blach=0
        laptopimage1=''
        if numbera==16:
            break
        column='Apartment Match '+str(numbera)
        numbera=numbera+1
        if imagecount==3:
            break
        print imagecount
        from openpyxl import Workbook
        from openpyxl import load_workbook

        wb = load_workbook('C:\Users\Kye Andreopoulus\Documents\Side Project\Master Tracker Test.xlsx')
        ws1 = wb.get_sheet_by_name("Sheet")
        try:
            print  dffinal4[column]['ApartmentAddress']
        except:
            break

        for row in ws1['B1':'B1000']:
            for cell in row:
                if results2.loc[number]['First Name']+results2.loc[number]['Last Name'] in str(cell.value):
                    found=cell.row
                    for x in range(3,6):
                        try:
                            if dffinal4[column]['ApartmentAddress'] in ws1.cell(row=found, column=x).value:
                                print 'nope'
                                tobreak=1
                                break
                        except:
                            break
                if tobreak==1:
                    break

            if tobreak==1:
                break
                
        if tobreak==1:
            continue
        
        websitecheck='https://maps.googleapis.com/maps/api/streetview/metadata?size=600x600&location='+dffinal4[column]['ApartmentAddress']+'&fov=90&heading=235&pitch=10&key=AIzaSyCouL7oS-mK8sfx_TDdQQ8GfoU7WDeNRuQ'


        for s in dffinal4[column]['ApartmentAddress'].split():
            if s.isdigit():
                print 'path1'
                blach=1
                site=requests.get(websitecheck)
                if 'ZERO_RESULTS' in site.json():
                    print 'no results'
                    if 'craigslist' in dffinal4[column]['Website']:
                        link=requests.get(dffinal4[column]['Website'])
                        soup=bs4.BeautifulSoup(link.text,'lxml')
                        listing=soup.find_all('a', {'class':'thumb'})
                        for x in listing: 
                            laptopimage1= x['href']
                            break
                    elif 'kijiji' in dffinal4[column]['Website']:
                        link=requests.get(dffinal4[column]['Website'])
                        soup=bs4.BeautifulSoup(link.text,'lxml')
                        listing2=soup.find_all('img')
                        for x in listing2:
                            if '$_35' in x['src']:
                                laptopimage1= x['src']
                                break
                    import urllib
                    desktopimage1='C:\Users\Kye Andreopoulus\Pictures\\'+laptopimage1[-15:]
                    print laptopimage1
                    try:
                        urllib.urlretrieve(laptopimage1, desktopimage1)
                        fp = open(desktopimage1, 'rb')
                        msgImage = MIMEImage(fp.read())
                        fp.close()
                        imagecount=imagecount+1
                        if locationlist!='':
                            locationlist=locationlist+'&markers=label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('google map, Canada',' ')
                            secondnumber=secondnumber+1
                            goodmatch.append(column)
                        else:
                            locationlist='label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('google map, Canada',' ')
                            secondnumber=secondnumber+1
                            goodmatch.append(column)
                        break
                    except:
                        continue
                    
                    
                
                

                if imagecount==0:
                
                    image1='''<img src="https://maps.googleapis.com/maps/api/streetview?size=600x300&location=&pitch=8&key=AIzaSyAbJyx_DmAR5uWGrp6FYCOvh2PW9jhHWkw" alt="Mountain View" style="width:600px;height:300px;">'''
                
                    insert1=image1.rfind('&pitch')
                    address = dffinal4[column]['ApartmentAddress']
                    image1=image1[:insert1]+address+image1[insert1:]
                    imagecount=imagecount+1
                    print image1
                    if locationlist!='':
                        locationlist=locationlist+'&markers=label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('google map, Canada',' ')
                        secondnumber=secondnumber+1
                        goodmatch.append(column)
                        break
                    else:
                        locationlist='label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('google map, Canada',' ')
                        secondnumber=secondnumber+1
                        goodmatch.append(column)
                    continue

                    
                
                if imagecount==1:
                    nextimage='''<img src="https://maps.googleapis.com/maps/api/streetview?size=600x300&location=&pitch=8&key=AIzaSyAbJyx_DmAR5uWGrp6FYCOvh2PW9jhHWkw" alt="Mountain View" style="width:600px;height:300px;">'''
                    insert1=nextimage.rfind('&pitch')
                    address = dffinal4[column]['ApartmentAddress']
                    nextimage=nextimage[:insert1]+address+nextimage[insert1:]
                    imagecount=imagecount+1
                    print nextimage
                    if locationlist!='':
                        locationlist=locationlist+'&markers=label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('google map, Canada',' ')
                        secondnumber=secondnumber+1
                        goodmatch.append(column)
                        break
                    else:
                        locationlist='label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('google map, Canada',' ')
                        secondnumber=secondnumber+1
                        goodmatch.append(column)
                    continue
                    
                if imagecount==2:
                    nextnextimage='''<img src="https://maps.googleapis.com/maps/api/streetview?size=600x300&location=&pitch=8&key=AIzaSyAbJyx_DmAR5uWGrp6FYCOvh2PW9jhHWkw" alt="Mountain View" style="width:600px;height:300px;">'''
                    insert1=nextnextimage.rfind('&pitch')
                    address = dffinal4[column]['ApartmentAddress']
                    nextnextimage=nextnextimage[:insert1]+address+nextnextimage[insert1:]
                    imagecount=imagecount+1
                    print nextimage
                    if locationlist!='':
                        locationlist=locationlist+'&markers=label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('google map, Canada',' ')
                        secondnumber=secondnumber+1
                        goodmatch.append(column)
                        break
                    else:
                        locationlist='label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('google map, Canada',' ')
                        secondnumber=secondnumber+1
                        goodmatch.append(column)
                    continue
                
                

        
        else:
            pass
        if blach==1:
            continue
        try:
            print 'path2'
            print dffinal4[column]['Website']
            if 'craigslist' in dffinal4[column]['Website']:
                link=requests.get(dffinal4[column]['Website'])
                soup=bs4.BeautifulSoup(link.text,'lxml')
                listing=soup.find_all('a', {'class':'thumb'})
                for x in listing: 
                    laptopimage1= x['href']
                    break
            elif 'kijiji' in dffinal4[column]['Website']:
                link=requests.get(dffinal4[column]['Website'])
                soup=bs4.BeautifulSoup(link.text,'lxml')
                listing2=soup.find_all('img')
                for x in listing2:
                    if '$_35' in x['src']:
                        laptopimage1= x['src']
                        break
            import urllib
            desktopimage1='C:\Users\Kye Andreopoulus\Pictures\\'+laptopimage1.replace('/','')[-15:]
            print laptopimage1

            urllib.urlretrieve(laptopimage1, desktopimage1)
            fp = open(desktopimage1, 'rb')
            if imagecount==0:
                msgImage = MIMEImage(fp.read())
                fp.close()
                img1='1st img'
            elif imagecount==1:
                msgImage2 = MIMEImage(fp.read())
                fp.close()
                img2= '2nd img'
            elif imagecount==2:
                msgImage3 = MIMEImage(fp.read())
                fp.close()
                img3= '3rd img'
            print 'this one'
            imagecount=imagecount+1
            #except:
             #   continue
            print imagecount
            if locationlist!='':
                locationlist=locationlist+'&markers=label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('googlemap',' ')
                secondnumber=secondnumber+1
                goodmatch.append(column)
            else:
                locationlist='label:'+str(secondnumber)+'%7C'+dffinal4[column]['ApartmentAddress'].replace('&',' ').replace('googlemap',' ')
                secondnumber=secondnumber+1
                goodmatch.append(column)
        except:
            continue
            
            
            
    print goodmatch   
        
    text1='Dear '+results2.loc[number]['First Name']+','
    text2='Thank you for trying Apartment Match from apartmentrank.com. Check out a few great matches (links below) we have hand-picked for you. These apartments were matched to you based on your answers to our lifestyle questionnaire as well as reviews of the apartments, popularity and other key factors.'
    text3='Enjoy!'    
        
    print locationlist
    
    
    image2='''<center><img src="https://gallery.mailchimp.com/1f5c0f61a871710fd71feb50a/images/290156af-a82d-48ab-89e3-7285552627b7.jpeg"style="width:700px;height:200px;"></center>'''
    image3='''<img src="https://maps.googleapis.com/maps/api/staticmap?zoom=11&size=500x300&maptype=roadmap&markers=&key=AIzaSyCouL7oS-mK8sfx_TDdQQ8GfoU7WDeNRuQ">'''
    insert2=image3.rfind('&key')
    address2 = locationlist
    image3=image3[:insert2]+address2+image3[insert2:]
    image3=str(image3)
    
    
    try:
        hyperlink1='''<a href='''+dffinal4[goodmatch[0]]['Website']+'''>Check out the comparison below, then click here to see this apartment.'''+'''</a>'''
    except: 
        pass
    try:
        hyperlink2='''<a href='''+dffinal4[goodmatch[1]]['Website']+'''>Check out the comparison below, then click here to see this apartment.'''+'''</a>'''
    except:
        pass
    try:    
        hyperlink3='''<a href='''+dffinal4[goodmatch[2]]['Website']+'''>Check out the comparison below, then click here to see this apartment.'''+'''</a>'''
    except:
        pass
    
    try:
        if img1=='':
            match1='''<b>'''+dffinal4[goodmatch[0]]['Apartment Price Range (including utilities, parking)']+'''</b><br>Distance from work: <b>'''+dffinal4[goodmatch[0]]['What is the closest intersection to work?']+'''<br></b>Distance from Hangout: <b>'''+dffinal4[goodmatch[0]]['Outside of work, what part of the city are you usually in? ']+'''</b><br>Apartment Type: <b>'''+dffinal4[goodmatch[0]]['How Many Bedrooms Were You Looking For?']+'''<br></b>Basement Or Above-Ground: <b>'''+dffinal4[goodmatch[0]]['Basement or Above Ground']+'''</b><br>'''
            try:
                match1=match1+'''Allows Pets:<b>'''+ dffinal4[goodmatch[0]]['Do you have any pets?']+ '''</b><br>'''+str(hyperlink1)+'''<br><br>'''+image1
            except:
                match1=match1+str(hyperlink1)+'<br><br><br>'+image1
        else:
            match1='''<b>'''+dffinal4[goodmatch[0]]['Apartment Price Range (including utilities, parking)']+'''</b><br>Distance from work: <b>'''+dffinal4[goodmatch[0]]['What is the closest intersection to work?']+'''</b><br>Distance from Hangout: <b>'''+dffinal4[goodmatch[0]]['Outside of work, what part of the city are you usually in? ']+'''</b><br>Apartment Type: <b>'''+dffinal4[goodmatch[0]]['How Many Bedrooms Were You Looking For?']+'''</b><br>Basement Or Above-Ground: <b>'''+dffinal4[goodmatch[0]]['Basement or Above Ground']+'''</b><br>'''
            try:
                match1=match1+'''Allows Pets: <b>''' +dffinal4[goodmatch[0]]['Do you have any pets?']+ '''</b><br>'''+str(hyperlink1)+'<br>'+'''<br>'''+'''<img src="cid:image10"style="width:500px;height:300px;">'''
            except:
                match1=match1+ str(hyperlink1)+'<br><br>'+'''<img src="cid:image10"style="width:500px;height:300px;">'''

        if img2=='':
            match2='''<b>'''+dffinal4[goodmatch[1]]['Apartment Price Range (including utilities, parking)']+'''</b><br>Distance from work: <b>'''+dffinal4[goodmatch[1]]['What is the closest intersection to work?']+'''</b><br>Distance from Hangout: <b>'''+dffinal4[goodmatch[1]]['Outside of work, what part of the city are you usually in? ']+'''</b><br>Apartment Type: <b>'''+dffinal4[goodmatch[1]]['How Many Bedrooms Were You Looking For?']+'''</b><br>Basement Or Above-Ground: <b>'''+dffinal4[goodmatch[1]]['Basement or Above Ground']+'''</b><br>'''
            try:
                match2=match2+'''Allows Pets: <b>''' +dffinal4[goodmatch[1]]['Do you have any pets?']+'</b><br>'+ str(hyperlink2)+'<br>'+'''<br>'''+nextimage
            except:
                match2=match2+ str(hyperlink2)+'<br>'+'''<br>'''+nextimage
        else:
            match2='''<b>'''+dffinal4[goodmatch[1]]['Apartment Price Range (including utilities, parking)']+'''</b><br>Distance from work: <b>'''+dffinal4[goodmatch[1]]['What is the closest intersection to work?']+'''</b><br>Distance from Hangout: <b>'''+dffinal4[goodmatch[1]]['Outside of work, what part of the city are you usually in? ']+'''</b><br>Apartment Type: <b>'''+dffinal4[goodmatch[1]]['How Many Bedrooms Were You Looking For?']+'''</b><br>Basement Or Above-Ground: <b>'''+dffinal4[goodmatch[1]]['Basement or Above Ground']+'''</b><br>'''
            try:
                match2=match2+'''Allows Pets: <b>''' +dffinal4[goodmatch[1]]['Do you have any pets?']+ '''</b><br>'''+ str(hyperlink2)+'<br>'+'''<br>'''+'''<img src="cid:image11"style="width:500px;height:300px;">'''
            except:
                match2=match2+ str(hyperlink2)+'<br>'+'''<br>'''+'''<img src="cid:image11"style="width:500px;height:300px;">'''

        if img3=='':
            match3='''<b>'''+dffinal4[goodmatch[2]]['Apartment Price Range (including utilities, parking)']+'''</b><br>Distance from work: <b>'''+dffinal4[goodmatch[2]]['What is the closest intersection to work?']+'''</b><br>Distance from Hangout: <b>'''+dffinal4[goodmatch[2]]['Outside of work, what part of the city are you usually in? ']+'''</b><br>Apartment Type: <b>'''+dffinal4[goodmatch[2]]['How Many Bedrooms Were You Looking For?']+'''</b><br>Basement Or Above-Ground: <b>'''+dffinal4[goodmatch[2]]['Basement or Above Ground']+'''</b><br>'''
            try:
                match3=match3+'''Allows Pets: <b>''' +dffinal4[goodmatch[2]]['Do you have any pets?']+'''</b><br>'''+ str(hyperlink3)+'<br>'+'''<br>'''+nextnextimage
            except:
                match3=match3+str(hyperlink3)+'<br>'+'''<br>'''+nextnextimage
        else:
            match3='''<b>'''+dffinal4[goodmatch[2]]['Apartment Price Range (including utilities, parking)']+'''</b><br>Distance from work: <b>'''+dffinal4[goodmatch[2]]['What is the closest intersection to work?']+'''</b><br>Distance from Hangout: <b>'''+dffinal4[goodmatch[2]]['Outside of work, what part of the city are you usually in? ']+'''</b><br>Apartment Type: <b>'''+dffinal4[goodmatch[2]]['How Many Bedrooms Were You Looking For?']+'''</b><br>Basement Or Above-Ground: <b>'''+dffinal4[goodmatch[2]]['Basement or Above Ground']+'''</b><br>'''
            try:
                match3=match3+'''Allows Pets: <b>''' +dffinal4[goodmatch[2]]['Do you have any pets?']+'''</b><br>'''+ str(hyperlink3)+'<br>'+'''<br>'''+'''<img src="cid:image12"style="width:500px;height:300px;">'''
            except:
                match3=match3+str(hyperlink3)+'<br>'+ '''<br>'''+'''<img src="cid:image12"style="width:500px;height:300px;">'''
    except:
        pass
        
    testlist=['Gym Within 10 Mins','Pharmacy Within 10 Mins','Restaurant Within 10 Mins', 'Grocery Within 10 Mins', 'Park Within 10 Mins','Coffee Shop Within 10 Mins']
    neighbourhoodlist=['What age are you, in spirit?','For whom?','Average Income of Neighbourhood']
    
    if match1=='':
        HEADER = '''
        <html>
            <head>

            </head>
            <body>
        '''
        FOOTER = '''
            </body>
        </html>
        '''

        data=HEADER + dffake.to_html(classes='df').encode('utf8') + FOOTER

        import smtplib

        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        # me == my email address
        # you == recipient's email address
        me = "torontoapartmentrank@gmail.com"
        you = "torontoapartmentrank@gmail.com "

        # Create message container - the correct MIME type is multipart/alternative.
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'No Results for ' +results2['First Name'].iloc[number]
        msg['From'] = me
        msg['To'] = you


        # Create the body of the message (a plain-text and an HTML version).
        text = "Hi!\nHow are you?\nHere is the link you wanted:\nhttp://www.python.org"
        html = data

        # Record the MIME types of both parts - text/plain and text/html.
        part1 = MIMEText(text, 'plain')
        part2 = MIMEText(html, 'html')

        # Attach parts into message container.
        # According to RFC 2046, the last part of a multipart message, in this case
        # the HTML message, is best and preferred.
        msg.attach(part1)
        msg.attach(part2)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login("torontoapartmentrank@gmail.com", "Milou511!")

        server.sendmail("torontoapartmentrank@gmail.com", "torontoapartmentrank@gmail.com", msg.as_string())
        # sendmail function takes 3 arguments: sender's addrwalkess, recipient's address
        # and message to send - here it is sent as one string.
        server.quit()
    
    elif match2=='':
        dfnearby=dffinal4[[goodmatch[0]]].ix[testlist]
        dfneighbour=dffinal4[[goodmatch[0]]].ix[neighbourhoodlist]
    elif match3=='':
        dfnearby=dffinal4[[goodmatch[0],goodmatch[1]]].ix[testlist]
        dfneighbour=dffinal4[[goodmatch[0],goodmatch[1]]].ix[neighbourhoodlist]
    else:
        dfnearby=dffinal4[[goodmatch[0],goodmatch[1],goodmatch[2]]].ix[testlist]
        dfneighbour=dffinal4[[goodmatch[0],goodmatch[1],goodmatch[2]]].ix[neighbourhoodlist]
        
        
    columnlist=[]
    for x in range(0,len(dfnearby.columns)):
        y=x+1
        columnlist.append('Apartment Match ' +str(y))
        
    dfnearby.columns=columnlist
    dfneighbour.columns=columnlist
    dfneighbour=dfneighbour.transpose()
    dfneighbour.columns=['Average Age of Postal Code', 'Demographics of Postal Code', 'Average Income of Postal Code']
    dfneighbour=dfneighbour.transpose()
     
    
    if match2=='':
        stringlist=(  str(dffinal4[goodmatch[0]]['ApartmentAddress'])+'''</b><br>'''+str(match1))
                    
    elif match3=='':
        stringlist=( str(dffinal4[goodmatch[0]]['ApartmentAddress'])+'''</b><br>'''+str(match1)+'<b><br><br><br><br>'+
                     str(dffinal4[goodmatch[1]]['ApartmentAddress'])+'''</b><br>'''+str(match2))
    else:
        stringlist=(str(dffinal4[goodmatch[0]]['ApartmentAddress'])+'''</b><br>'''+str(match1)+'<b><br><br><br><br>'+
                    str(dffinal4[goodmatch[1]]['ApartmentAddress'])+'''</b><br>'''+str(match2)+'<b><br><br><br><br>'
                    +str(dffinal4[goodmatch[2]]['ApartmentAddress'])+'''</b><br>'''+str(match3))
          
        
    
    
    
    data=(HEADER + image2+'''<br />'''+'''<br />'''+'''<br />'''
          #+'''<?phprename(image7,"pictures");?>'''
          +text1+'''<br><br><br>'''+text2+'''<br><br>'''
          +text3+'''<br />'''+'''<br />'''+'''<br /><b><br>'''+
          stringlist
          
          
          
          
          +'''<br><br><br>'''+'''<br />'''+'''<br />'''+'<br><br><b> Map: </b><br><br>'+image3+'<br><br>'+
          '<b> Mostly Singles or Families: </b> '+'''<br><br>'''+dfneighbour.to_html(classes='df').encode('utf8')+'''<br>'''
          +'''<br>'''+'''<br>'''+"<b> Top Ranked Amenities According to Yelp, Within Walking Distance: </b> "+'''<br><br>'''
          +dfnearby.to_html(classes='df').encode('utf8')
          +'''<br>'''
          
          +FOOTER)

    import smtplib

    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.MIMEImage import MIMEImage


    # me == my email address
    # you == recipient's email address
    me = "torontoapartmentrank@gmail.com"
    you = results2['Email to send you your matches'].iloc[number]

    # Create message container - the correct MIME type is multipart/alternative.
    msg = MIMEMultipart('alternative')
    msg['Subject'] = "Your Apartment Match Results For Today"
    msg['From'] = me
    msg['To'] = you


    # Create the body of the message (a plain-text and an HTML version).
    text = "Your perfect apartment matches"
    html = data

    # Record the MIME types of both parts - text/plain and text/html.
    part1 = MIMEText(text, 'plain')
    part2 = MIMEText(html, 'html')
    #part3=MIMEText('<b>Some <i>HTML</i> text</b> and an image.<br><img src="cid:image10"><br>Nifty!', 'html')

    # Attach parts into message container.
    # According to RFC 2046, the last part of a multipart message, in this case
    # the HTML message, is best and preferred.
    msg.attach(part1)
    msg.attach(part2)
    #msg.attach(part3)
    
    msgAlternative = MIMEMultipart('alternative')
    msg.attach(msgAlternative)
    
    # We reference the image in the IMG SRC attribute by the ID we give it below
    msgText = MIMEText(html, 'html')
    msgAlternative.attach(msgText)   
    
    # Define the image's ID as referenced above
    if img1 != '':
        msgImage.add_header('Content-ID', '<image10>')
        msg.attach(msgImage) 
    
    if img2 != '':
        msgImage2.add_header('Content-ID', '<image11>')
        msg.attach(msgImage2) 
        
    if img3 != '':
        msgImage3.add_header('Content-ID', '<image12>')
        msg.attach(msgImage3) 

    

    import imaplib
    import time
    conn = imaplib.IMAP4_SSL('imap.gmail.com')
    conn.login('torontoapartmentrank@gmail.com', 'Milou511!')
    conn.select('[Gmail]/Drafts')
    import email
    conn.append("[Gmail]/Drafts",
                '',
                imaplib.Time2Internaldate(time.time()),
                str(email.message_from_string(msg.as_string())))
    
    print 'Sent to Draft'
    # sendmail function takes 3 arguments: sender's address, recipient's address
    # and message to send - here it is sent as one string.
    
   
    import pandas
    from openpyxl import Workbook
    from openpyxl import load_workbook

    wb = load_workbook('C:\Users\Kye Andreopoulus\Documents\Side Project\Master Tracker Test.xlsx')
    ws1 = wb.get_sheet_by_name("Sheet")

    rownumber=ws1.get_highest_row()
    ts = time.time()

    ws1.cell(row=rownumber+1, column=1).value=datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
    ws1.cell(row=rownumber+1, column=2).value=results2.loc[number]['First Name']+results2.loc[number]['Last Name']
    try:
        ws1.cell(row=rownumber+1, column=3).value=dffinal4[goodmatch[0]]['ApartmentAddress']
        ws1.cell(row=rownumber+1, column=4).value=dffinal4[goodmatch[1]]['ApartmentAddress']
        ws1.cell(row=rownumber+1, column=5).value=dffinal4[goodmatch[2]]['ApartmentAddress']
    except:
        pass 

    wb.save('C:\Users\Kye Andreopoulus\Documents\Side Project\Master Tracker Test.xlsx')
    
    driver.quit()
    driver2.quit()
    driver4.quit()
    driver13.quit()
    driver14.quit()
    driver15.quit()
    driver16.quit()
    driver17.quit()
    driver18.quit()
    driver19.quit()
    #CODE TO SEND ALL E-MAILS FROM DRAFT

    for x in range (0,100):
        try:
            import imaplib
            import time
            import smtplib 
            from datetime import datetime, timedelta
            conn = imaplib.IMAP4_SSL('imap.gmail.com')
            conn.login('torontoapartmentrank@gmail.com', 'Milou511!')
            conn.select('[Gmail]/Drafts')
            today = datetime.today()
            cutoff = today - timedelta(days=1)
            dt = cutoff.strftime('%d-%b-%Y')
            print dt
            result, data =conn.search(None, '(SINCE %s)' %(dt,))
            ids = data[0] # data is a list.
            id_list = ids.split() # ids is a space separated string
            latest_email_id = id_list[-x] # get the latest

            result, data = conn.fetch(latest_email_id, "(RFC822)") # fetch the email body (RFC822)             for the given ID

            raw_email = data[0][1]

            towhom= raw_email.find('To')
            stop=raw_email[towhom+4:].find('com')
            if stop>100:
                stop=raw_email[towhom+4:].find('ca')
            if stop>100:
                stop=raw_email[towhom+4:].find('net')

            server = smtplib.SMTP('smtp.bananatag.com', 587)
            server.starttls()
            server.login("torontoapartmentrank@gmail.com", "Milou511!") 
            server.sendmail("torontoapartmentrank@gmail.com",raw_email[towhom+4:towhom+stop+4+3] , raw_email)
            server.quit()
            conn.store(3, '+FLAGS', '\\Deleted')
            conn.expunge()
        except:
            print 'Done email'
            break